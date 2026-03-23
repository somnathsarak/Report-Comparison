import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import shutil


class ExcelReportComparator:
    """
    A class to compare Excel files from different releases and create a comparison report.
    
    Attributes:
        project_folder (str): Path to the project folder
        release1_path (str): Path to Release1 folder
        release2_path (str): Path to Release2 folder
        comparison_path (str): Path to comparison output folder
    """
    
    def __init__(self, project_folder):
        """
        Initialize the ExcelReportComparator.
        
        Args:
            project_folder (str): Path to the project folder
        """
        self.project_folder = project_folder
        self.release1_path = os.path.join(project_folder, 'Release1')
        self.release2_path = os.path.join(project_folder, 'Release2')
        self.comparison_path = os.path.join(project_folder, 'comparison')
        
    def create_folder_structure(self):
        """
        Create the necessary folder structure if it doesn't exist.
        """
        print("Creating folder structure...")
        os.makedirs(os.path.join(self.release1_path, 'A42'), exist_ok=True)
        os.makedirs(os.path.join(self.release2_path, 'A42'), exist_ok=True)
        os.makedirs(os.path.join(self.comparison_path, 'A42'), exist_ok=True)
        print("Folder structure created successfully.")
    
    def read_excel_file(self, file_path):
        """
        Read Excel file and return as DataFrame.
        
        Args:
            file_path (str): Path to the Excel file
            
        Returns:
            pd.DataFrame: DataFrame containing the Excel data
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            df = pd.read_excel(file_path, sheet_name=0)  # Read first sheet
            print(f"Successfully read: {file_path}")
            return df
        except Exception as e:
            print(f"Error reading file {file_path}: {str(e)}")
            raise
    
    def create_comparison_report(self, release1_file, release2_file, output_file):
        """
        Create a comparison report with separate sheets for Release1 and Release2 data.
        
        Args:
            release1_file (str): Path to Release1 Excel file
            release2_file (str): Path to Release2 Excel file
            output_file (str): Path to save the comparison report
        """
        try:
            # Read both Excel files
            print("\nReading Excel files...")
            df_release1 = self.read_excel_file(release1_file)
            df_release2 = self.read_excel_file(release2_file)
            
            # Create a new workbook
            print("Creating comparison report...")
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            # Add Release1 data to sheet
            ws1 = wb.create_sheet(title="Release1_Report")
            for r_idx, row in enumerate(dataframe_to_rows(df_release1, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws1.cell(row=r_idx, column=c_idx, value=value)
            
            # Add Release2 data to sheet
            ws2 = wb.create_sheet(title="Release2_Report")
            for r_idx, row in enumerate(dataframe_to_rows(df_release2, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws2.cell(row=r_idx, column=c_idx, value=value)
            
            # Save the workbook
            os.makedirs(os.path.dirname(output_file), exist_ok=True)
            wb.save(output_file)
            print(f"Comparison report saved to: {output_file}")
            print(f"\nReport Details:")
            print(f"  - Release1_Report sheet: {len(df_release1)} rows, {len(df_release1.columns)} columns")
            print(f"  - Release2_Report sheet: {len(df_release2)} rows, {len(df_release2.columns)} columns")
            
        except Exception as e:
            print(f"Error creating comparison report: {str(e)}")
            raise
    
    def compare_reports(self, folder_name='A42'):
        """
        Main method to compare reports from Release1 and Release2.
        
        Args:
            folder_name (str): Folder name containing the Excel files (default: 'A42')
        """
        try:
            # Define file paths
            release1_file = os.path.join(self.release1_path, folder_name, 'excelFile1.xlsx')
            release2_file = os.path.join(self.release2_path, folder_name, 'excelFile2.xlsx')
            output_file = os.path.join(self.comparison_path, folder_name, 'comparison_report.xlsx')
            
            # Check if source files exist
            if not os.path.exists(release1_file):
                print(f"Warning: Release1 file not found at {release1_file}")
                print("Please ensure excelFile1.xlsx exists in Release1/A42/ folder")
                return False
            
            if not os.path.exists(release2_file):
                print(f"Warning: Release2 file not found at {release2_file}")
                print("Please ensure excelFile2.xlsx exists in Release2/A42/ folder")
                return False
            
            # Create comparison report
            self.create_comparison_report(release1_file, release2_file, output_file)
            return True
            
        except Exception as e:
            print(f"Error during comparison: {str(e)}")
            return False


def main():
    """
    Main function to run the Excel report comparison.
    """
    # Get the current directory or specify your project folder
    project_folder = os.path.dirname(os.path.abspath(__file__))
    
    # Create comparator instance
    comparator = ExcelReportComparator(project_folder)
    
    # Create folder structure
    comparator.create_folder_structure()
    
    # Run comparison
    success = comparator.compare_reports(folder_name='A42')
    
    if success:
        print("\n\u2713 Comparison completed successfully!")
    else:
        print("\n\u2717 Comparison failed. Please check the file paths and ensure files exist.")


if __name__ == "__main__":
    main()
