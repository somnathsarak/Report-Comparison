import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os


class AllowanceReportCombiner:
    """
    Combines Allowance Delivery Reports from different releases into a single Excel file
    with separate sheets for each release.
    """

    def __init__(self, output_path='comparison_report.xlsx'):
        """
        Initialize the combiner.

        Args:
            output_path (str): Path where the combined report will be saved
        """
        self.output_path = output_path
        self.workbook = None

    def read_excel_all_sheets(self, file_path):
        """
        Read all sheets from an Excel file and return as a dictionary.

        Args:
            file_path (str): Path to the Excel file

        Returns:
            dict: Dictionary with sheet names as keys and DataFrames as values
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            excel_file = pd.ExcelFile(file_path)
            sheets_dict = {}
            for sheet_name in excel_file.sheet_names:
                sheets_dict[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"Successfully read {len(sheets_dict)} sheet(s) from: {file_path}")
            return sheets_dict
        except Exception as e:
            print(f"Error reading file {file_path}: {str(e)}")
            raise

    def combine_reports(self, release1_path, release2_path, sheet1_name='POST_R3.6_A42_CA_QC_JOINT_AUCTION',
                       sheet2_name='R3.7_A42_CA_QC_JOINT_AUCTION'):
        """
        Combine two Excel reports into a single file with renamed sheets.

        Args:
            release1_path (str): Path to Release 1 (R3.6) Excel file
            release2_path (str): Path to Release 2 (R3.7) Excel file
            sheet1_name (str): Name for the first sheet (R3.6 data)
            sheet2_name (str): Name for the second sheet (R3.7 data)
        """
        try:
            # Create new workbook
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)  # Remove default sheet

            # Read Release 1 (R3.6) data
            print(f"\nReading Release 1 (R3.6) file: {release1_path}")
            release1_sheets = self.read_excel_all_sheets(release1_path)

            # Read Release 2 (R3.7) data
            print(f"Reading Release 2 (R3.7) file: {release2_path}")
            release2_sheets = self.read_excel_all_sheets(release2_path)

            # Get the first sheet from each file (assuming data is in the first sheet)
            release1_data = list(release1_sheets.values())[0]
            release2_data = list(release2_sheets.values())[0]

            # Add Release 1 data to new workbook with custom sheet name
            print(f"\nAdding Release 1 data to sheet: {sheet1_name}")
            ws1 = self.workbook.create_sheet(title=sheet1_name)
            for r_idx, row in enumerate(dataframe_to_rows(release1_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws1.cell(row=r_idx, column=c_idx, value=value)

            # Add Release 2 data to new workbook with custom sheet name
            print(f"Adding Release 2 data to sheet: {sheet2_name}")
            ws2 = self.workbook.create_sheet(title=sheet2_name)
            for r_idx, row in enumerate(dataframe_to_rows(release2_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws2.cell(row=r_idx, column=c_idx, value=value)

            # Save the combined workbook
            os.makedirs(os.path.dirname(self.output_path) or '.', exist_ok=True)
            self.workbook.save(self.output_path)
            print(f"\n✓ Combined report saved to: {self.output_path}")

            # Print summary
            print(f"\nReport Summary:")
            print(f"  Sheet 1 ({sheet1_name}): {len(release1_data)} rows, {len(release1_data.columns)} columns")
            print(f"  Sheet 2 ({sheet2_name}): {len(release2_data)} rows, {len(release2_data.columns)} columns")
            print(f"  Total sheets in output: 2")

            return True

        except Exception as e:
            print(f"Error combining reports: {str(e)}")
            return False


def main():
    """
    Main function to combine allowance reports.
    """
    # Define file paths
    release1_file = 'Release1/A42/Allowance-Delivery-Report-POST_R3.6_A42_CA_QC_JOINT_AUCTION.xlsx'
    release2_file = 'Release2/A42/Allowance-Delivery-Report-R3.7_A42_CA_QC_JOINT_AUCTION.xlsx'
    output_file = 'comparison/A42/allowance_comparison_report.xlsx'

    # Create combiner instance
    combiner = AllowanceReportCombiner(output_path=output_file)

    # Combine the reports
    success = combiner.combine_reports(
        release1_path=release1_file,
        release2_path=release2_file,
        sheet1_name='POST_R3.6_A42_CA_QC_JOINT_AUCTION',
        sheet2_name='R3.7_A42_CA_QC_JOINT_AUCTION'
    )

    if success:
        print("\n✓ Report comparison completed successfully!")
    else:
        print("\n✗ Report comparison failed.")


if __name__ == "__main__":
    main()
