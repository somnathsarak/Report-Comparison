import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys


class AllowanceReportCombiner:
    """
    Combines Allowance Delivery Reports from different releases into a single Excel file
    with separate sheets for each release.
    
    This script is designed to be run from the comparison/A42/ directory and combines
    reports from Release1/A42/ and Release2/A42/ directories.
    """

    def __init__(self, output_path='allowance_comparison_report.xlsx'):
        """
        Initialize the combiner.

        Args:
            output_path (str): Path where the combined report will be saved
        """
        self.output_path = output_path
        self.workbook = None
        self.base_dir = os.path.dirname(os.path.abspath(__file__))
        self.project_dir = os.path.dirname(os.path.dirname(self.base_dir))  # Navigate to ProjectFolder

    def read_excel_all_sheets(self, file_path):
        """
        Read all sheets from an Excel file and return as a dictionary.

        Args:
            file_path (str): Path to the Excel file

        Returns:
            dict: Dictionary with sheet names as keys and DataFrames as values
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            excel_file = pd.ExcelFile(file_path)
            sheets_dict = {}
            for sheet_name in excel_file.sheet_names:
                sheets_dict[sheet_name] = pd.read_excel(file_path, sheet_name=sheet_name)
            print(f"Successfully read {len(sheets_dict)} sheet(s) from: {os.path.basename(file_path)}")
            return sheets_dict
        except Exception as e:
            print(f"Error reading file {file_path}: {str(e)}")
            raise

    def combine_reports(self, release1_filename, release2_filename, 
                       sheet1_name='POST_R3.6_A42_CA_QC_JOINT_AUCTION',
                       sheet2_name='R3.7_A42_CA_QC_JOINT_AUCTION'):
        """
        Combine two Excel reports from Release1/A42/ and Release2/A42/ into a single file.

        Args:
            release1_filename (str): Filename of Release 1 (R3.6) Excel file
            release2_filename (str): Filename of Release 2 (R3.7) Excel file
            sheet1_name (str): Name for the first sheet (R3.6 data)
            sheet2_name (str): Name for the second sheet (R3.7 data)
        """
        try:
            # Construct full paths relative to project folder
            release1_path = os.path.join(self.project_dir, 'Release1', 'A42', release1_filename)
            release2_path = os.path.join(self.project_dir, 'Release2', 'A42', release2_filename)

            print("=" * 70)
            print("ALLOWANCE DELIVERY REPORT COMBINER")
            print("=" * 70)
            print(f"\nProject Directory: {self.project_dir}")
            print(f"Output Directory: {os.path.dirname(self.output_path) or 'current'}\n")

            # Create new workbook
            self.workbook = Workbook()
            self.workbook.remove(self.workbook.active)  # Remove default sheet

            # Read Release 1 (R3.6) data
            print(f"Reading Release 1 (R3.6) file...")
            print(f"  Path: {release1_path}")
            release1_sheets = self.read_excel_all_sheets(release1_path)

            # Read Release 2 (R3.7) data
            print(f"\nReading Release 2 (R3.7) file...")
            print(f"  Path: {release2_path}")
            release2_sheets = self.read_excel_all_sheets(release2_path)

            # Get the first sheet from each file (assuming data is in the first sheet)
            release1_data = list(release1_sheets.values())[0]
            release2_data = list(release2_sheets.values())[0]

            # Add Release 1 data to new workbook with custom sheet name
            print(f"\nCreating sheets in output workbook...")
            print(f"  Sheet 1: {sheet1_name}")
            ws1 = self.workbook.create_sheet(title=sheet1_name)
            for r_idx, row in enumerate(dataframe_to_rows(release1_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws1.cell(row=r_idx, column=c_idx, value=value)

            # Add Release 2 data to new workbook with custom sheet name
            print(f"  Sheet 2: {sheet2_name}")
            ws2 = self.workbook.create_sheet(title=sheet2_name)
            for r_idx, row in enumerate(dataframe_to_rows(release2_data, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    ws2.cell(row=r_idx, column=c_idx, value=value)

            # Save the combined workbook
            os.makedirs(os.path.dirname(self.output_path) or '.', exist_ok=True)
            self.workbook.save(self.output_path)
            print(f"\n" + "=" * 70)
            print(f"SUCCESS: Combined report saved to: {os.path.basename(self.output_path)}")
            print("=" * 70)

            # Print summary
            print(f"\nReport Summary:")
            print(f"  Sheet 1 ({sheet1_name}): {len(release1_data)} rows, {len(release1_data.columns)} columns")
            print(f"  Sheet 2 ({sheet2_name}): {len(release2_data)} rows, {len(release2_data.columns)} columns")
            print(f"  Total sheets in output: 2")
            print(f"\nOutput file location: {os.path.abspath(self.output_path)}")

            return True

        except FileNotFoundError as e:
            print(f"\n" + "=" * 70)
            print(f"ERROR: {str(e)}")
            print("=" * 70)
            print(f"\nPlease ensure the following files exist:")
            print(f"  1. Release1/A42/{release1_filename}")
            print(f"  2. Release2/A42/{release2_filename}")
            return False
        except Exception as e:
            print(f"\n" + "=" * 70)
            print(f"ERROR: {str(e)}")
            print("=" * 70)
            return False


def main():
    """
    Main function to combine allowance reports.
    Usage: python A42_combine_allowance_reports.py [R36_filename] [R37_filename]
    """
    # Define file names (can be overridden via command line arguments)
    if len(sys.argv) > 2:
        release1_file = sys.argv[1]
        release2_file = sys.argv[2]
    else:
        # Default file names - update these to match your actual file names
        release1_file = 'Allowance-Delivery-Report-POST_R3.6_A42_CA_QC_JOINT_AUCTION-07-01-02-01-07-2026.xlsx'
        release2_file = 'Allowance-Delivery-Report-R3.7_A42_CA_QC_JOINT_AUCTION-6March_01-03-06-2026-2.xlsx'

    output_file = 'allowance_comparison_report.xlsx'

    # Create combiner instance
    combiner = AllowanceReportCombiner(output_path=output_file)

    # Combine the reports
    success = combiner.combine_reports(
        release1_filename=release1_file,
        release2_filename=release2_file,
        sheet1_name='POST_R3.6_A42_CA_QC_JOINT_AUCTION',
        sheet2_name='R3.7_A42_CA_QC_JOINT_AUCTION'
    )

    if success:
        print("\n✓ Report comparison completed successfully!\n")
        sys.exit(0)
    else:
        print("\n✗ Report comparison failed.\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
